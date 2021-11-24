import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from boltons.setutils import IndexedSet

product_list_url = "https://light-home.pl/2-lampy?n=500"

categories = set()

wb = Workbook()
ws = wb.active

response = requests.get(product_list_url)

if(response.status_code!=200):
    print("Error - status code: "+response.status_code)
    exit(0)

soup1 = BeautifulSoup(response.text, 'html.parser')

i=1

#Nagłówki
ws['A1'] = "ID"
ws['B1'] = "Nazwa"
ws['C1'] = "Adresy URL zdjęcia (x,y,z...)"
ws['D1'] = "Opis"
ws['E1'] = "Cecha(Nazwa:Wartość:Pozycja:Indywidualne)"
ws['F1'] = "Cena zawiera podatek. (brutto)"
ws['G1'] = "ID reguły podatku"
ws['H1'] = "Kategorie (x,y,z...)"
ws['I1'] = "Marka"
ws['J1'] = "Indeks #"
ws['K1'] = "Podsumowanie"
ws['L1'] = "Ilość"

product_links = soup1.find_all('a', class_="product-name")
for link in product_links:
    product_response = requests.get(link.get('href'))
    print(product_response.status_code)
    soup2 = BeautifulSoup(product_response.text, 'html.parser')
    print(link.get('href'))
    
    i += 1
    
    ws['A'+str(i)] = i
    
    #Zdjecia
    img_links = soup2.find_all("div", class_="bigpic_item")
    images = ""
    for img_link in img_links:
        tmp = img_link.find("a").get('href')
        print(tmp)
        images += tmp + ','
    ws['C'+str(i)] = images
    
    #Nazwa
    tmp = str(soup2.find("h1",class_="kp-prodtitle").get_text())
    print(tmp)
    ws['B'+str(i)] = tmp
    
    #Informacje o produkcie
    tmp = str(soup2.find(id="idTab1").find("div",class_="pa_content").find("div").find("p"))
    print(tmp)
    ws['D'+str(i)] = tmp
    
    #Dane techniczne
    tmp = soup2.find(id="techData").find("table").find_all("tr")
    tmp_str=""
    i2 = 1
    for x in tmp:
        tmp_str+=str(x).replace(' class="odd"','').replace(' class="even"','').replace(",","/").replace("</td><td>",":").replace("</td></tr>",":"+str(i2)+",").replace("<tr><td>","")
        i2 += 1
    print(tmp_str)
    ws['E'+str(i)] = tmp_str
    
    #Cena
    tmp = str(soup2.find(id="our_price_display").get_text())
    print(tmp)
    ws['F'+str(i)] = tmp
    
    #Id podatku
    ws['G'+str(i)] = "1"
    
    #Kategoria
    tmp = str(soup2.find(id="breadcrumb_wrapper").find("ul").get_text().replace("Jesteś tu: ","").replace(">",",")).replace(','+str(ws['B']),"")
    pos_end = tmp.rfind(',')
    pos_start = tmp.rfind(',',0,pos_end)+1
    categories.add(tmp[:pos_end])
    print(tmp[pos_start:pos_end])
    ws['H'+str(i)] = tmp[pos_start:pos_end]
    
    #Marka
    tmp = soup2.find(id="product_manufacturer_info")
    if tmp != None:
        tmp=str(tmp.find("a").get_text())
        print(tmp[1:])
        ws['I'+str(i)] = tmp[1:]
    
    #Kod producenta
    tmp = str(soup2.find(id="product_reference").find("span").get_text())
    print(tmp)
    ws['J'+str(i)] = tmp
    ws['K'+str(i)] = tmp
    
    #Początkowy stan magazynowy
    ws['L'+str(i)] = "10"
    
#wb.save('import_products.xlsx')
    
#przetworzenie kategorii do formatu importu
wb_categories = Workbook()
ws_categories = wb_categories.active

categories_processed = IndexedSet()

for category in categories:
    tmp = category.split(",")
    last_cat = ""
    i = 0
    for x in tmp:
        if last_cat != "":
            i += 1
            categories_processed.add((last_cat,x,i))
        last_cat = x

print(categories_processed)

categories_processed.sort(key=lambda tup: tup[2])  # sortowanie według głębokości (trzeci element tuple)

ws_categories['A1'] = "ID"
ws_categories['B1'] = "Kategoria nadrzędna"
ws_categories['C1'] = "Nazwa"

ws_categories['A2'] = 3
ws_categories['B2'] = "Kategoria główna"
ws_categories['C2'] = "Sklep z lampami"

i = 3
for category in categories_processed:
    if category[0]=="" or category[1]=="":
        continue
    ws_categories['A'+str(i)] = i+1
    ws_categories['B'+str(i)] = category[0]
    ws_categories['C'+str(i)] = category[1]
    i += 1

wb_categories.save("import_categories.xlsx")


#zamiana kategorii w imporcie produktów na ID
for row_cat in ws_categories.rows:
    print(row_cat[2].value)
    for row_prod in ws.rows:
        if row_prod[7].value == row_cat[2].value:
            row_prod[7].value = row_cat[0].value
        
wb.save("import_products.xlsx")
